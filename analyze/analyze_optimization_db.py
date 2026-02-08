"""
Optimization Results Analyzer v3 (SQLite DB version)

Analyzes SQLite databases with corrected tsl_target=0 data to find best
parameter combinations for live trading, grouped by IB parameters.

Analysis modes:
1. IB + Buffer: Group by (ib_start, ib_end, ib_tz, ib_wait, ib_buffer_pct)
2. IB + Buffer + MaxDist: Group by (ib_start, ib_end, ib_tz, ib_wait, ib_buffer_pct, max_distance_pct)

Output:
- Excel files with multiple sheets for different grouping modes
- Python dicts ready to copy to strategy_logic.py
"""

import sqlite3
import pandas as pd
import numpy as np
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Configuration
ANALYZE_DIR = Path(__file__).parent
MAX_DRAWDOWN_FILTER = 12.0  # Max allowed drawdown in R
TOP_N = 25  # Number of top groups to show

# Data sources (SQLite DBs)
DB_PATHS = {
    "GER40": ANALYZE_DIR / "GER40_optimization.db",
    "XAUUSD": ANALYZE_DIR / "XAUUSD_optimization.db",
}

# Variation prefixes in columns
VARIATIONS = ["ocae", "tcwe", "reverse", "rev_rb"]


def load_data_from_db(symbol: str) -> pd.DataFrame:
    """Load data from SQLite database."""
    db_path = DB_PATHS[symbol]
    print(f"[INFO] Loading {symbol} from {db_path.name}...")

    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM results", conn)
    conn.close()

    print(f"  -> Loaded {len(df):,} rows")
    return df


def parse_params(df: pd.DataFrame) -> pd.DataFrame:
    """Parse params_json column into separate columns."""
    print("[INFO] Parsing params_json...")

    def safe_parse(x):
        if isinstance(x, dict):
            return x
        elif isinstance(x, str):
            return json.loads(x)
        return {}

    params_list = df["params_json"].apply(safe_parse).tolist()
    params_df = pd.DataFrame(params_list)
    result = pd.concat([df.drop(columns=["params_json"]), params_df], axis=1)
    return result


def get_group_key_ib_buffer(row) -> tuple:
    """Group key: IB params + Buffer (same for all variations)."""
    return (
        row["ib_start"],
        row["ib_end"],
        row["ib_timezone"],
        int(row["ib_wait_minutes"]),
        float(row["ib_buffer_pct"]),
    )


def get_group_key_ib_buffer_maxdist(row) -> tuple:
    """Group key: IB params + Buffer + MaxDist (same for all variations)."""
    return (
        row["ib_start"],
        row["ib_end"],
        row["ib_timezone"],
        int(row["ib_wait_minutes"]),
        float(row["ib_buffer_pct"]),
        float(row["max_distance_pct"]),
    )


def find_best_variation_params(group_df: pd.DataFrame, variation: str) -> dict:
    """Find best parameters for a variation within a group."""
    prefix = variation

    total_r_col = f"{prefix}_total_r"
    max_dd_col = f"{prefix}_max_drawdown"
    sharpe_col = f"{prefix}_sharpe_ratio"
    winrate_col = f"{prefix}_winrate"
    trades_col = f"{prefix}_trades"

    # Filter by max drawdown
    filtered = group_df[group_df[max_dd_col] <= MAX_DRAWDOWN_FILTER].copy()

    if filtered.empty:
        filtered = group_df.copy()

    # Filter by TSL constraint: TSL_SL < RR_TARGET + 1
    # This ensures TSL can actually execute on live bot (price must reach target before SL adjustment)
    tsl_target_col = f"{prefix}_tsl_target" if f"{prefix}_tsl_target" in filtered.columns else "tsl_target"
    tsl_sl_col = f"{prefix}_tsl_sl" if f"{prefix}_tsl_sl" in filtered.columns else "tsl_sl"
    rr_col = f"{prefix}_rr_target" if f"{prefix}_rr_target" in filtered.columns else "rr_target"

    # Apply TSL constraint: TSL disabled (tsl_target=0) OR tsl_sl < rr_target + 1
    if tsl_sl_col in filtered.columns and rr_col in filtered.columns:
        valid_tsl_mask = (filtered[tsl_target_col] <= 0) | (filtered[tsl_sl_col] < filtered[rr_col] + 1)
        filtered = filtered[valid_tsl_mask]

    if filtered.empty:
        return None

    best_idx = filtered[total_r_col].idxmax()
    best_row = filtered.loc[best_idx]

    result = {
        "variation": variation.upper(),
        "total_r": float(best_row[total_r_col]),
        "sharpe": float(best_row[sharpe_col]) if pd.notna(best_row[sharpe_col]) else 0,
        "winrate": float(best_row[winrate_col]) if pd.notna(best_row[winrate_col]) else 0,
        "trades": int(best_row[trades_col]) if pd.notna(best_row[trades_col]) else 0,
        "max_drawdown": float(best_row[max_dd_col]) if pd.notna(best_row[max_dd_col]) else 0,
        # Common params
        "ib_start": best_row["ib_start"],
        "ib_end": best_row["ib_end"],
        "ib_timezone": best_row["ib_timezone"],
        "ib_wait_minutes": int(best_row["ib_wait_minutes"]),
        "ib_buffer_pct": float(best_row["ib_buffer_pct"]),
        "max_distance_pct": float(best_row["max_distance_pct"]),
        # Variation-specific params
        "trade_window_minutes": int(best_row["trade_window_minutes"]),
        "rr_target": float(best_row["rr_target"]),
        "stop_mode": best_row["stop_mode"],
        "tsl_target": float(best_row["tsl_target"]),
        "tsl_sl": float(best_row["tsl_sl"]),
        "min_sl_pct": float(best_row["min_sl_pct"]),
        "rev_rb_enabled": bool(best_row.get("rev_rb_enabled", False)),
        "rev_rb_pct": float(best_row.get("rev_rb_pct", 1.0)),
    }

    return result


def analyze_with_grouping(df: pd.DataFrame, group_key_func, mode_name: str) -> list:
    """Analyze data with specified grouping function."""
    print(f"\n[INFO] Analyzing with grouping: {mode_name}")

    df["group_key"] = df.apply(group_key_func, axis=1)
    groups = df["group_key"].unique()
    print(f"  -> Found {len(groups)} unique groups")

    group_results = []

    for group_key in groups:
        group_df = df[df["group_key"] == group_key]

        variation_results = {}
        combined_total_r = 0
        total_trades = 0
        weighted_sharpe_sum = 0

        for var in VARIATIONS:
            best = find_best_variation_params(group_df, var)
            if best:
                variation_results[var] = best
                combined_total_r += best["total_r"]
                trades = best["trades"]
                if trades > 0:
                    weighted_sharpe_sum += best["sharpe"] * trades
                    total_trades += trades

        if not variation_results:
            continue

        weighted_avg_sharpe = weighted_sharpe_sum / total_trades if total_trades > 0 else 0

        first_var = list(variation_results.values())[0]

        result = {
            "group_key": group_key,
            "ib_start": first_var["ib_start"],
            "ib_end": first_var["ib_end"],
            "ib_timezone": first_var["ib_timezone"],
            "ib_wait": first_var["ib_wait_minutes"],
            "ib_buffer_pct": first_var["ib_buffer_pct"],
            "max_distance_pct": first_var["max_distance_pct"],
            "combined_total_r": combined_total_r,
            "weighted_sharpe": weighted_avg_sharpe,
            "total_trades": total_trades,
            "variations": variation_results,
            "num_combinations": len(group_df),
        }

        group_results.append(result)

    group_results.sort(key=lambda x: x["combined_total_r"], reverse=True)

    print(f"\n[INFO] Top {TOP_N} groups ({mode_name}):")
    for i, g in enumerate(group_results[:TOP_N], 1):
        print(f"  {i}. IB {g['ib_start']}-{g['ib_end']} Wait {g['ib_wait']}m Buffer {g['ib_buffer_pct']*100:.0f}%", end="")
        if "maxdist" in mode_name.lower():
            print(f" MaxDist {g['max_distance_pct']*100:.0f}%", end="")
        print(f" -> R: {g['combined_total_r']:.2f}, wSharpe: {g['weighted_sharpe']:.2f}, Trades: {g['total_trades']}")

    return group_results[:TOP_N]


def analyze_symbol(symbol: str) -> dict:
    """Analyze optimization results for a symbol with multiple grouping modes."""
    print(f"\n{'='*60}")
    print(f"ANALYZING {symbol}")
    print(f"{'='*60}")

    df = load_data_from_db(symbol)
    df = parse_params(df)

    top_ib_buffer = analyze_with_grouping(df, get_group_key_ib_buffer, "IB + Buffer")
    top_ib_buffer_maxdist = analyze_with_grouping(df, get_group_key_ib_buffer_maxdist, "IB + Buffer + MaxDist")

    return {
        "symbol": symbol,
        "total_combinations": len(df),
        "top_ib_buffer": top_ib_buffer,
        "top_ib_buffer_maxdist": top_ib_buffer_maxdist,
    }


def create_excel_report(analysis: dict, output_path: Path):
    """Create Excel report with multiple sheets."""
    print(f"\n[INFO] Creating Excel report: {output_path}")

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_fill_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    # Sheet 1: Summary (IB + Buffer)
    ws1 = wb.active
    ws1.title = "IB+Buffer Summary"

    ws1.append([f"Optimization Analysis - {analysis['symbol']} (DB v3)"])
    ws1.append([f"Total Combinations: {analysis['total_combinations']:,}"])
    ws1.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws1.append([])
    ws1.append(["MODE: IB + Buffer (same buffer for all variations)"])
    ws1.append([])

    headers = ["Rank", "IB Window", "TZ", "Wait", "Buffer %", "Combined R", "wSharpe", "Trades"]
    ws1.append(headers)
    for cell in ws1[ws1.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    for i, g in enumerate(analysis["top_ib_buffer"], 1):
        ws1.append([
            i,
            f"{g['ib_start']}-{g['ib_end']}",
            g['ib_timezone'],
            g['ib_wait'],
            f"{g['ib_buffer_pct']*100:.0f}%",
            round(g['combined_total_r'], 2),
            round(g['weighted_sharpe'], 2),
            g['total_trades'],
        ])

    # Sheet 2: Variation Details (IB + Buffer)
    ws2 = wb.create_sheet("IB+Buffer Details")

    headers = [
        "Rank", "IB Window", "Wait", "Buffer %", "Variation",
        "Total R", "Sharpe", "WR %", "Trades", "MaxDD",
        "Window", "RR", "Stop", "TSL_T", "TSL_SL", "MaxDist %"
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    for i, g in enumerate(analysis["top_ib_buffer"], 1):
        for var_name, v in g["variations"].items():
            ws2.append([
                i,
                f"{g['ib_start']}-{g['ib_end']}",
                g['ib_wait'],
                f"{g['ib_buffer_pct']*100:.0f}%",
                var_name.upper(),
                round(v["total_r"], 2),
                round(v["sharpe"], 2),
                round(v["winrate"], 1),
                v["trades"],
                round(v["max_drawdown"], 1),
                v["trade_window_minutes"],
                v["rr_target"],
                v["stop_mode"],
                v["tsl_target"],
                v["tsl_sl"],
                f"{v['max_distance_pct']*100:.0f}%",
            ])

    # Sheet 3: Summary (IB + Buffer + MaxDist)
    ws3 = wb.create_sheet("IB+Buffer+MaxDist Summary")

    ws3.append([f"Optimization Analysis - {analysis['symbol']} (DB v3)"])
    ws3.append([f"Total Combinations: {analysis['total_combinations']:,}"])
    ws3.append([])
    ws3.append(["MODE: IB + Buffer + MaxDist (same buffer AND maxdist for all variations)"])
    ws3.append([])

    headers = ["Rank", "IB Window", "TZ", "Wait", "Buffer %", "MaxDist %", "Combined R", "wSharpe", "Trades"]
    ws3.append(headers)
    for cell in ws3[ws3.max_row]:
        cell.font = header_font
        cell.fill = header_fill_green

    for i, g in enumerate(analysis["top_ib_buffer_maxdist"], 1):
        ws3.append([
            i,
            f"{g['ib_start']}-{g['ib_end']}",
            g['ib_timezone'],
            g['ib_wait'],
            f"{g['ib_buffer_pct']*100:.0f}%",
            f"{g['max_distance_pct']*100:.0f}%",
            round(g['combined_total_r'], 2),
            round(g['weighted_sharpe'], 2),
            g['total_trades'],
        ])

    # Sheet 4: Variation Details (IB + Buffer + MaxDist)
    ws4 = wb.create_sheet("IB+Buffer+MaxDist Details")

    headers = [
        "Rank", "IB Window", "Wait", "Buffer %", "MaxDist %", "Variation",
        "Total R", "Sharpe", "WR %", "Trades", "MaxDD",
        "Window", "RR", "Stop", "TSL_T", "TSL_SL"
    ]
    ws4.append(headers)
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill_green

    for i, g in enumerate(analysis["top_ib_buffer_maxdist"], 1):
        for var_name, v in g["variations"].items():
            ws4.append([
                i,
                f"{g['ib_start']}-{g['ib_end']}",
                g['ib_wait'],
                f"{g['ib_buffer_pct']*100:.0f}%",
                f"{g['max_distance_pct']*100:.0f}%",
                var_name.upper(),
                round(v["total_r"], 2),
                round(v["sharpe"], 2),
                round(v["winrate"], 1),
                v["trades"],
                round(v["max_drawdown"], 1),
                v["trade_window_minutes"],
                v["rr_target"],
                v["stop_mode"],
                v["tsl_target"],
                v["tsl_sl"],
            ])

    # Auto-width columns
    for ws in [ws1, ws2, ws3, ws4]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    wb.save(output_path)
    print(f"  -> Saved: {output_path}")


def generate_params_dict(analysis: dict, mode: str = "ib_buffer") -> str:
    """Generate Python dict for strategy_logic.py."""
    symbol = analysis["symbol"]

    if mode == "ib_buffer":
        top_group = analysis["top_ib_buffer"][0]
        version = "V8"
        mode_desc = "IB + Buffer"
    else:
        top_group = analysis["top_ib_buffer_maxdist"][0]
        version = "V8_STRICT"
        mode_desc = "IB + Buffer + MaxDist"

    min_sl = 0.0015 if symbol == "GER40" else 0.001

    lines = [
        f"# {symbol} {version} parameters - DB Optimizer ({datetime.now().strftime('%Y-%m-%d')})",
        f"# Mode: {mode_desc} (same for all variations)",
        f"# Best: IB {top_group['ib_start']}-{top_group['ib_end']} ({top_group['ib_timezone']}) Wait {top_group['ib_wait']}m",
        f"# Buffer: {top_group['ib_buffer_pct']*100:.0f}%, MaxDist: {top_group['max_distance_pct']*100:.0f}%",
        f"# Combined Total R: {top_group['combined_total_r']:.2f}, Weighted Sharpe: {top_group['weighted_sharpe']:.2f}, Total Trades: {top_group['total_trades']}",
        f"{symbol}_PARAMS_{version} = {{",
    ]

    for var_name in ["REV_RB", "Reverse", "TCWE", "OCAE"]:
        var_key = var_name.lower()
        if var_key not in top_group["variations"]:
            continue

        v = top_group["variations"][var_key]

        lines.append(f'    "{var_name}": {{')
        lines.append(f'        "IB_START": "{v["ib_start"]}",')
        lines.append(f'        "IB_END": "{v["ib_end"]}",')
        lines.append(f'        "IB_TZ": "{v["ib_timezone"]}",')
        lines.append(f'        "IB_WAIT": {v["ib_wait_minutes"]},')
        lines.append(f'        "TRADE_WINDOW": {v["trade_window_minutes"]},')
        lines.append(f'        "RR_TARGET": {v["rr_target"]},')
        lines.append(f'        "STOP_MODE": "{v["stop_mode"]}",')
        lines.append(f'        "TSL_TARGET": {v["tsl_target"]},')
        lines.append(f'        "TSL_SL": {v["tsl_sl"]},')
        lines.append(f'        "MIN_SL_PCT": {min_sl},')

        if var_name == "REV_RB":
            lines.append(f'        "REV_RB_PCT": {v["rev_rb_pct"]},')
            lines.append(f'        "REV_RB_ENABLED": {str(v["rev_rb_enabled"])},')
        else:
            lines.append(f'        "REV_RB_ENABLED": False,')

        lines.append(f'        "IB_BUFFER_PCT": {v["ib_buffer_pct"]},')
        lines.append(f'        "MAX_DISTANCE_PCT": {v["max_distance_pct"]},')
        lines.append(f'        # R: {v["total_r"]:.2f}, Sharpe: {v["sharpe"]:.2f}, WR: {v["winrate"]:.1f}%, Trades: {v["trades"]}, MaxDD: {v["max_drawdown"]:.1f}')
        lines.append('    },')

    lines.append("}")

    return "\n".join(lines)


def main():
    print("="*60)
    print("OPTIMIZATION RESULTS ANALYZER v3 (SQLite DB)")
    print("(Groups by IB+Buffer and IB+Buffer+MaxDist)")
    print("="*60)

    for symbol in ["GER40", "XAUUSD"]:
        try:
            analysis = analyze_symbol(symbol)

            # Create Excel report
            excel_path = ANALYZE_DIR / f"optimization_analysis_{symbol}_v3.xlsx"
            create_excel_report(analysis, excel_path)

            # Generate Python params dict - IB + Buffer mode
            params_code = generate_params_dict(analysis, mode="ib_buffer")
            params_path = ANALYZE_DIR / f"best_params_{symbol}_V8.py"
            params_path.write_text(params_code, encoding="utf-8")
            print(f"[INFO] Saved params dict: {params_path}")

            # Generate Python params dict - IB + Buffer + MaxDist mode
            params_code_strict = generate_params_dict(analysis, mode="ib_buffer_maxdist")
            params_path_strict = ANALYZE_DIR / f"best_params_{symbol}_V8_STRICT.py"
            params_path_strict.write_text(params_code_strict, encoding="utf-8")
            print(f"[INFO] Saved params dict: {params_path_strict}")

            # Print both params for convenience
            print(f"\n{'-'*60}")
            print(f"BEST PARAMS FOR {symbol} (IB + Buffer):")
            print(f"{'-'*60}")
            print(params_code)

            print(f"\n{'-'*60}")
            print(f"BEST PARAMS FOR {symbol} (IB + Buffer + MaxDist):")
            print(f"{'-'*60}")
            print(params_code_strict)

        except Exception as e:
            print(f"[ERROR] Failed to analyze {symbol}: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "="*60)
    print("ANALYSIS COMPLETE")
    print("="*60)


if __name__ == "__main__":
    main()
