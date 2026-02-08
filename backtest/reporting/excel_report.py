"""
Excel Report Generator

Creates detailed Excel reports with:
- "Trades" sheet: Individual trade details
- "Statistics" sheet: Performance metrics
- "By Variation" sheet: Breakdown by strategy variation
- Conditional formatting for P/L columns
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..emulator.models import TradeLog

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generates Excel reports for backtest results.

    Creates professional-looking Excel files with:
    - Detailed trade log
    - Performance statistics
    - Conditional formatting
    - Auto-column width
    """

    def __init__(self):
        """Initialize Excel report generator."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, Excel reports will be disabled")

        # Styles
        self.header_font = Font(bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") if OPENPYXL_AVAILABLE else None
        self.header_alignment = Alignment(horizontal="center", vertical="center") if OPENPYXL_AVAILABLE else None
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ) if OPENPYXL_AVAILABLE else None

    def generate_report(
        self,
        trade_log: List[TradeLog],
        metrics_report,
        output_path: Path,
        symbol: str,
        initial_balance: float,
        ib_data: Optional[Dict[str, Dict[str, float]]] = None,
    ) -> Optional[Path]:
        """
        Generate Excel report.

        Args:
            trade_log: List of TradeLog entries
            metrics_report: MetricsReport with calculated metrics
            output_path: Path to save Excel file
            symbol: Trading symbol
            initial_balance: Starting balance
            ib_data: Optional dict of date -> {ibh, ibl, eq} values

        Returns:
            Path to saved file or None if failed
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Cannot generate Excel report - openpyxl not installed")
            return None

        try:
            wb = Workbook()

            # Create sheets
            self._create_trades_sheet(wb, trade_log, symbol, ib_data)
            self._create_statistics_sheet(wb, metrics_report, initial_balance)
            self._create_variation_sheet(wb, metrics_report)

            # Remove default sheet if exists
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Save
            excel_path = output_path / "results.xlsx"
            wb.save(excel_path)
            logger.info(f"Excel report saved to {excel_path}")
            return excel_path

        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
            return None

    def _create_trades_sheet(
        self,
        wb: Workbook,
        trade_log: List[TradeLog],
        symbol: str,
        ib_data: Optional[Dict[str, Dict[str, float]]] = None,
    ) -> None:
        """Create Trades sheet with detailed trade information."""
        ws = wb.create_sheet("Trades", 0)

        # Prepare data
        rows = []
        for i, trade in enumerate(trade_log, 1):
            if not trade.exit_time:
                continue  # Skip open trades

            # Get IB data for this trade's date
            ibh, ibl, eq = None, None, None
            if ib_data and trade.entry_time:
                date_key = trade.entry_time.strftime("%Y-%m-%d")
                if date_key in ib_data:
                    ibh = ib_data[date_key].get("ibh")
                    ibl = ib_data[date_key].get("ibl")
                    eq = ib_data[date_key].get("eq")

            # Calculate R value (in price units, not USD)
            initial_risk = abs(trade.entry_price - trade.sl) if trade.sl else 0
            if initial_risk > 0:
                # R = price movement / initial risk
                if trade.direction and trade.direction.lower() == "long":
                    price_movement = trade.exit_price - trade.entry_price
                else:  # short
                    price_movement = trade.entry_price - trade.exit_price
                r_value = price_movement / initial_risk
            else:
                r_value = 0

            row = {
                "#": i,
                "Date": trade.entry_time.strftime("%Y-%m-%d") if trade.entry_time else "",
                "Symbol": symbol,
                "Variation": trade.variation or "",
                "Direction": trade.direction.upper() if trade.direction else "",
                "IBH": ibh,
                "IBL": ibl,
                "EQ": eq,
                "Entry Time": trade.entry_time.strftime("%H:%M:%S") if trade.entry_time else "",
                "Entry Price": trade.entry_price,
                "Stop Loss": trade.sl,
                "Take Profit": trade.tp,
                "Exit Time": trade.exit_time.strftime("%H:%M:%S") if trade.exit_time else "",
                "Exit Price": trade.exit_price,
                "Exit Reason": trade.exit_reason or "",
                "Volume": trade.volume,
                "P/L ($)": trade.profit,
                "R": round(r_value, 2),
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Format numbers
                if headers[col_num - 1] in ["Entry Price", "Exit Price", "Stop Loss", "Take Profit", "IBH", "IBL", "EQ"]:
                    cell.number_format = "#,##0.00"
                elif headers[col_num - 1] == "P/L ($)":
                    cell.number_format = "#,##0.00"
                elif headers[col_num - 1] == "R":
                    cell.number_format = "0.00"
                elif headers[col_num - 1] == "Volume":
                    cell.number_format = "0.00"

        # Apply borders
        for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, max_col=len(headers)):
            for cell in row:
                cell.border = self.thin_border

        # Auto column width
        for col_num, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row_num in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(1, col_num).column_letter].width = min(max_length + 2, 20)

        # Conditional formatting for P/L column
        pl_col = headers.index("P/L ($)") + 1
        pl_col_letter = ws.cell(1, pl_col).column_letter

        # Green for positive, red for negative
        ws.conditional_formatting.add(
            f"{pl_col_letter}2:{pl_col_letter}{len(rows) + 1}",
            ColorScaleRule(
                start_type="num", start_value=-500, start_color="FF6666",
                mid_type="num", mid_value=0, mid_color="FFFF99",
                end_type="num", end_value=500, end_color="66FF66"
            )
        )

        # Conditional formatting for R column
        r_col = headers.index("R") + 1
        r_col_letter = ws.cell(1, r_col).column_letter

        ws.conditional_formatting.add(
            f"{r_col_letter}2:{r_col_letter}{len(rows) + 1}",
            ColorScaleRule(
                start_type="num", start_value=-2, start_color="FF0000",
                mid_type="num", mid_value=0, mid_color="FFFF00",
                end_type="num", end_value=2, end_color="00FF00"
            )
        )

        # Alternating row colors
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_num in range(2, len(rows) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, len(headers) + 1):
                    # Don't override conditional formatting
                    if col_num not in [pl_col, r_col]:
                        ws.cell(row=row_num, column=col_num).fill = light_fill

    def _create_statistics_sheet(
        self,
        wb: Workbook,
        metrics_report,
        initial_balance: float,
    ) -> None:
        """Create Statistics sheet with performance metrics."""
        ws = wb.create_sheet("Statistics")

        stats = [
            ("TRADE STATISTICS", ""),
            ("Total Trades", metrics_report.total_trades),
            ("Winning Trades", metrics_report.winning_trades),
            ("Losing Trades", metrics_report.losing_trades),
            ("Breakeven Trades", metrics_report.breakeven_trades),
            ("Win Rate (%)", f"{metrics_report.win_rate:.1f}"),
            ("", ""),
            ("PROFIT METRICS", ""),
            ("Total Profit ($)", f"{metrics_report.total_profit:.2f}"),
            ("Gross Profit ($)", f"{metrics_report.gross_profit:.2f}"),
            ("Gross Loss ($)", f"{metrics_report.gross_loss:.2f}"),
            ("Profit Factor", f"{metrics_report.profit_factor:.2f}"),
            ("Average Win ($)", f"{metrics_report.avg_win:.2f}"),
            ("Average Loss ($)", f"{metrics_report.avg_loss:.2f}"),
            ("Average Trade ($)", f"{metrics_report.avg_trade:.2f}"),
            ("Win/Loss Ratio", f"{metrics_report.win_loss_ratio:.2f}"),
            ("", ""),
            ("RISK METRICS", ""),
            ("Max Drawdown ($)", f"{metrics_report.max_drawdown:.2f}"),
            ("Max Drawdown (%)", f"{metrics_report.max_drawdown_pct:.1f}"),
            ("Max DD Duration (days)", f"{metrics_report.max_drawdown_duration_days:.0f}"),
            ("Sharpe Ratio", f"{metrics_report.sharpe_ratio:.2f}"),
            ("Sortino Ratio", f"{metrics_report.sortino_ratio:.2f}"),
            ("Calmar Ratio", f"{metrics_report.calmar_ratio:.2f}"),
            ("", ""),
            ("RETURN METRICS", ""),
            ("Initial Balance ($)", f"{initial_balance:.2f}"),
            ("Final Balance ($)", f"{initial_balance + metrics_report.total_profit:.2f}"),
            ("Total Return (%)", f"{metrics_report.total_return_pct:.1f}"),
            ("Annualized Return (%)", f"{metrics_report.annualized_return_pct:.1f}"),
            ("", ""),
            ("STREAK ANALYSIS", ""),
            ("Max Consecutive Wins", metrics_report.max_consecutive_wins),
            ("Max Consecutive Losses", metrics_report.max_consecutive_losses),
        ]

        # Headers
        ws.cell(1, 1, "Metric").font = self.header_font
        ws.cell(1, 1).fill = self.header_fill
        ws.cell(1, 2, "Value").font = self.header_font
        ws.cell(1, 2).fill = self.header_fill

        # Data
        for row_num, (metric, value) in enumerate(stats, 2):
            cell_metric = ws.cell(row=row_num, column=1, value=metric)
            cell_value = ws.cell(row=row_num, column=2, value=value)

            # Section headers
            if metric and not value:
                cell_metric.font = Font(bold=True, color="366092")

            # Key metrics highlighting
            if metric in ["Win Rate (%)", "Total Profit ($)", "Profit Factor", "Max Drawdown (%)"]:
                cell_metric.font = Font(bold=True)
                cell_value.font = Font(bold=True, color="0066CC")

        # Borders and width
        for row in ws.iter_rows(min_row=1, max_row=len(stats) + 1, max_col=2):
            for cell in row:
                cell.border = self.thin_border

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_variation_sheet(self, wb: Workbook, metrics_report) -> None:
        """Create By Variation sheet with breakdown by strategy variation."""
        ws = wb.create_sheet("By Variation")

        if not metrics_report.by_variation:
            ws.cell(1, 1, "No variation data available")
            return

        # Headers
        headers = ["Variation", "Trades", "Wins", "Losses", "Win Rate (%)", "Total P/L ($)", "Avg P/L ($)"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        # Data
        row_num = 2
        for var, stats in metrics_report.by_variation.items():
            ws.cell(row=row_num, column=1, value=var)
            ws.cell(row=row_num, column=2, value=stats["total_trades"])
            ws.cell(row=row_num, column=3, value=stats["winning_trades"])
            ws.cell(row=row_num, column=4, value=stats["losing_trades"])
            ws.cell(row=row_num, column=5, value=f"{stats['win_rate']:.1f}")
            ws.cell(row=row_num, column=6, value=f"{stats['total_profit']:.2f}")
            ws.cell(row=row_num, column=7, value=f"{stats['avg_profit']:.2f}")
            row_num += 1

        # Borders and width
        for row in ws.iter_rows(min_row=1, max_row=row_num - 1, max_col=len(headers)):
            for cell in row:
                cell.border = self.thin_border

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col_num).column_letter].width = 15
