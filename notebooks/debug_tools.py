"""
Debug Tools for Strategy Analysis

Provides functions for:
1. Running single-day backtests
2. Generating trade charts inline
3. Exporting trades to Excel with comment columns
4. Comparing results with reference data
"""

import sys
import logging
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List, Tuple

import pandas as pd
import numpy as np
import pytz

# Add parent paths for imports
DUAL_V4_PATH = Path(__file__).parent.parent
sys.path.insert(0, str(DUAL_V4_PATH))

from backtest.emulator import mt5_emulator as mt5_emu
from backtest.emulator.models import TradeLog
from backtest.adapter import BacktestExecutor, create_mt5_patch_module
from backtest.config import BacktestConfig, SymbolConfig
from backtest.risk_manager import BacktestRiskManager
from backtest.data_processor.data_ingestor import DataIngestor

logger = logging.getLogger(__name__)


class SingleDayBacktest:
    """
    Run backtest for a single day and get detailed results.
    """

    def __init__(
        self,
        symbol: str = "GER40",
        initial_balance: float = 50000.0,
        risk_pct: Optional[float] = None,
        risk_amount: Optional[float] = None,
        max_margin_pct: float = 40.0,
        data_path: Optional[str] = None,
        timezone: Optional[str] = None,
    ):
        """
        Initialize single-day backtest.

        Args:
            symbol: Trading symbol (GER40, XAUUSD)
            initial_balance: Starting balance
            risk_pct: Risk percentage (1.0 = 1%)
            risk_amount: Fixed risk amount ($)
            max_margin_pct: Max margin percentage
            data_path: Path to M1 data folder
            timezone: Instrument timezone (auto-detected if None)
        """
        self.symbol = symbol
        self.initial_balance = initial_balance
        self.risk_pct = risk_pct
        self.risk_amount = risk_amount
        self.max_margin_pct = max_margin_pct

        # Auto-detect timezone based on symbol
        if timezone is None:
            if symbol == "XAUUSD":
                timezone = "Asia/Tokyo"
            else:  # GER40 and others default to Europe/Berlin
                timezone = "Europe/Berlin"
        self.timezone = timezone

        # Default data paths
        if data_path is None:
            if symbol == "GER40":
                data_path = str(DUAL_V4_PATH / "data" / "GER40 1m 01_01_2023-04_11_2025")
            elif symbol == "XAUUSD":
                data_path = str(DUAL_V4_PATH / "data" / "XAUUSD 1m 01_01_2023-04_11_2025")
        self.data_path = data_path

        # Will be initialized on first run
        self.emulator = None
        self.executor = None
        self.strategy = None
        self.risk_manager = None
        self.m1_data = None

    def _setup(self):
        """Initialize emulator and load data."""
        # Symbol configuration
        if self.symbol == "GER40":
            symbol_config = SymbolConfig(
                name="GER40",
                trade_contract_size=1.0,
                trade_tick_value=1.0,
                trade_tick_size=0.05,
                volume_min=0.01,
                volume_max=100.0,
                volume_step=0.01,
                spread_points=1.0,
                digits=2,
            )
        else:  # XAUUSD
            symbol_config = SymbolConfig(
                name="XAUUSD",
                trade_contract_size=100.0,
                trade_tick_value=1.0,
                trade_tick_size=0.01,
                volume_min=0.01,
                volume_max=100.0,
                volume_step=0.01,
                spread_points=0.30,
                digits=2,
            )

        # Create config (symbols must be a dict)
        config = BacktestConfig(
            initial_balance=self.initial_balance,
            leverage=100,
            symbols={self.symbol: symbol_config},
        )

        # Create emulator
        self.emulator = mt5_emu.MT5Emulator()
        self.emulator.configure(config)

        # Load data
        ingestor = DataIngestor(self.data_path)
        self.m1_data = ingestor.load_all()
        self.emulator.load_m1_data(self.symbol, self.m1_data)

        # Create executor
        self.executor = BacktestExecutor(self.emulator, config)
        self.executor.connect(12345, "", "")

        # Create risk manager
        self.risk_manager = BacktestRiskManager(
            self.emulator,
            risk_pct=self.risk_pct,
            risk_amount=self.risk_amount,
            max_margin_pct=self.max_margin_pct,
        )

        # Patch MT5 module
        import sys as _sys
        _sys.modules['MetaTrader5'] = create_mt5_patch_module(self.emulator)

        # Import strategy after patching (with reload to pick up changes)
        import importlib
        import src.strategies.ib_strategy as ib_strategy_module
        import src.utils.strategy_logic as strategy_logic_module
        importlib.reload(strategy_logic_module)
        importlib.reload(ib_strategy_module)
        from src.strategies.ib_strategy import IBStrategy
        from src.utils.strategy_logic import GER40_PARAMS_PROD, XAUUSD_PARAMS_PROD

        # Get params for symbol (PROD parameters)
        params = GER40_PARAMS_PROD if self.symbol == "GER40" else XAUUSD_PARAMS_PROD

        # Create strategy
        self.strategy = IBStrategy(
            symbol=self.symbol,
            params=params,
            executor=self.executor,
            magic_number=1001,
            strategy_label="Debug"
        )

    def run_day(self, date: datetime) -> Dict[str, Any]:
        """
        Run backtest for a single day.

        Args:
            date: Date to backtest (datetime or string 'YYYY-MM-DD')

        Returns:
            Dict with:
                - trade: TradeLog or None
                - ib_data: Dict with IBH, IBL, EQ
                - signals: List of detected signals
                - candles: DataFrame of day's candles
                - error: Error message if any
        """
        if isinstance(date, str):
            date = datetime.strptime(date, "%Y-%m-%d")

        if self.emulator is None:
            self._setup()

        # Reset state (don't call full reset() as it disconnects)
        self.emulator.current_time = None
        self.emulator._account.balance = self.initial_balance
        self.emulator._account.equity = self.initial_balance
        self.emulator._account.margin_free = self.initial_balance
        self.emulator._positions.clear()
        self.emulator._trade_log.clear()
        self.strategy.reset_daily_state()

        # Get timezone
        tz = pytz.timezone(self.timezone)

        # Define day boundaries
        day_start = tz.localize(datetime(date.year, date.month, date.day, 7, 0, 0))
        day_end = tz.localize(datetime(date.year, date.month, date.day, 22, 0, 0))

        # Convert to UTC
        day_start_utc = day_start.astimezone(pytz.utc)
        day_end_utc = day_end.astimezone(pytz.utc)

        # Filter candles for this day
        day_candles = self.m1_data[
            (self.m1_data["time"] >= day_start_utc) &
            (self.m1_data["time"] <= day_end_utc)
        ].copy()

        if day_candles.empty:
            return {
                "trade": None,
                "ib_data": None,
                "signals": [],
                "candles": day_candles,
                "error": f"No data for {date.strftime('%Y-%m-%d')}",
            }

        # Collect data
        signals_detected = []
        ib_data = None
        current_ticket = None
        magic_number = 1001

        # Iterate through candles
        for idx, candle in day_candles.iterrows():
            candle_time = candle["time"]
            if hasattr(candle_time, 'to_pydatetime'):
                candle_time = candle_time.to_pydatetime()

            # Advance emulator time
            self.emulator.set_time(candle_time)

            # Check for signal only if no open position
            positions = self.executor.get_open_positions()
            has_position = any(p.magic == magic_number for p in positions)

            if not has_position:
                # Check for signal
                signal = self.strategy.check_signal(candle_time)

                # Collect IB data after calculation
                if self.strategy.ibh is not None and ib_data is None:
                    ib_data = {
                        "ibh": self.strategy.ibh,
                        "ibl": self.strategy.ibl,
                        "eq": self.strategy.eq,
                    }

                # Handle signal (only record the first valid signal)
                if signal is not None and len(signals_detected) == 0:
                    signals_detected.append({
                        "time": candle_time,
                        "direction": signal.direction,
                        "variation": signal.variation,
                        "entry_price": signal.entry_price,
                        "stop_loss": signal.stop_loss,
                        "take_profit": signal.take_profit,
                    })

                    # Calculate position size
                    lots = self.risk_manager.calculate_position_size(
                        self.symbol,
                        signal.entry_price,
                        signal.stop_loss,
                    )

                    if lots > 0:
                        # Place order
                        result = self.executor.place_order(
                            self.symbol,
                            signal,
                            lots,
                            magic_number,
                        )
                        if result.get("success"):
                            current_ticket = result.get("ticket")
                            # CRITICAL: Change state to prevent further signal detection
                            self.strategy.state = "POSITION_OPEN"

            else:
                # Check SL/TP and update position state (TSL logic)
                for position in positions:
                    if position.magic == magic_number:
                        high = float(candle["high"])
                        low = float(candle["low"])

                        # Save SL before TSL update to detect if TSL moved it
                        sl_before_tsl = position.sl

                        # First, run TSL logic (may move SL up if virtual TP hit)
                        # Use high/low from candle for proper TP detection
                        simulated_tick = {
                            "bid": low if position.type == 0 else high,
                            "ask": high if position.type == 0 else low,
                        }
                        self.strategy.update_position_state(
                            position,
                            simulated_tick,
                            current_time_utc=candle_time,
                        )

                        # Re-fetch position after TSL update (SL may have changed)
                        updated_positions = self.executor.get_open_positions()
                        updated_position = None
                        for p in updated_positions:
                            if p.ticket == position.ticket:
                                updated_position = p
                                break

                        if updated_position is None:
                            # Position was closed by time window
                            continue

                        # Check SL hit - candle must CROSS the SL level
                        # For LONG: high >= SL (was above) AND low <= SL (dropped to SL)
                        # For SHORT: low <= SL (was below) AND high >= SL (rose to SL)
                        sl_hit = False
                        current_sl = updated_position.sl

                        if updated_position.type == 0:  # LONG
                            # SL hit only if candle crossed SL from above
                            if current_sl > 0 and high >= current_sl and low <= current_sl:
                                sl_hit = True
                                self.emulator.close_position_by_ticket(
                                    updated_position.ticket,
                                    price=current_sl,
                                    exit_reason="sl"
                                )
                        else:  # SHORT
                            # SL hit only if candle crossed SL from below
                            if current_sl > 0 and low <= current_sl and high >= current_sl:
                                sl_hit = True
                                self.emulator.close_position_by_ticket(
                                    updated_position.ticket,
                                    price=current_sl,
                                    exit_reason="sl"
                                )

                        if sl_hit:
                            # Don't clear tsl_state here - we need it for charting
                            pass

        # Get trade result
        trade_log = self.emulator.get_trade_log()
        trade = trade_log[0] if trade_log else None

        # Preserve TSL state for charting before it could be cleared
        tsl_state_copy = None
        if self.strategy.tsl_state:
            tsl_state_copy = dict(self.strategy.tsl_state)
            if "tsl_history" in tsl_state_copy:
                tsl_state_copy["tsl_history"] = list(tsl_state_copy["tsl_history"])

        return {
            "trade": trade,
            "ib_data": ib_data,
            "signals": signals_detected,
            "candles": day_candles,
            "tsl_state": tsl_state_copy,  # Include TSL state for charting
            "error": None,
        }

    def generate_chart(
        self,
        result: Dict[str, Any],
        figsize: Tuple[int, int] = (14, 8),
        show_levels: bool = True,
        hours_after_ib: float = 4.0,
    ):
        """
        Generate matplotlib chart for the day's result.

        Args:
            result: Result from run_day()
            figsize: Figure size
            show_levels: Show IBH/IBL/EQ levels
            hours_after_ib: Limit chart to N hours after IB end (default 3h)

        Returns:
            matplotlib Figure
        """
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.patches import Rectangle

        candles = result["candles"]
        trade = result["trade"]
        ib_data = result["ib_data"]
        signals = result["signals"]

        if candles.empty:
            print(f"No candles to plot: {result.get('error')}")
            return None

        # Convert times to local timezone
        tz = pytz.timezone(self.timezone)
        candles = candles.copy()
        candles["time_local"] = candles["time"].dt.tz_convert(tz)

        # Limit chart time range: IB start - 30min to IB end + hours_after_ib
        # IB period is typically 08:00-08:30 for GER40
        ib_start_hour, ib_start_min = map(int, self.strategy.ib_start.split(":"))
        ib_end_hour, ib_end_min = map(int, self.strategy.ib_end.split(":"))

        first_date = candles["time_local"].iloc[0].date()
        chart_start = tz.localize(datetime(first_date.year, first_date.month, first_date.day,
                                           ib_start_hour, ib_start_min)) - timedelta(minutes=30)
        chart_end = tz.localize(datetime(first_date.year, first_date.month, first_date.day,
                                         ib_end_hour, ib_end_min)) + timedelta(hours=hours_after_ib)

        # Filter candles to chart range
        candles = candles[(candles["time_local"] >= chart_start) &
                          (candles["time_local"] <= chart_end)]

        if candles.empty:
            print(f"No candles in chart range {chart_start.strftime('%H:%M')} - {chart_end.strftime('%H:%M')}")
            return None

        fig, ax = plt.subplots(figsize=figsize)

        # Plot candlesticks
        width = 0.0005  # Bar width in days

        for idx, row in candles.iterrows():
            time = mdates.date2num(row["time_local"])
            o, h, l, c = row["open"], row["high"], row["low"], row["close"]

            color = "green" if c >= o else "red"

            # Body
            body_bottom = min(o, c)
            body_height = abs(c - o)
            rect = Rectangle(
                (time - width/2, body_bottom),
                width, body_height,
                facecolor=color, edgecolor=color, alpha=0.8
            )
            ax.add_patch(rect)

            # Wicks
            ax.plot([time, time], [l, body_bottom], color=color, linewidth=0.5)
            ax.plot([time, time], [body_bottom + body_height, h], color=color, linewidth=0.5)

        # Chart end for TSL segments
        chart_end_num = mdates.date2num(chart_end)

        # Plot IB levels
        if ib_data and show_levels:
            ibh = ib_data["ibh"]
            ibl = ib_data["ibl"]
            eq = ib_data["eq"]

            # Calculate IB period boundaries for the rectangle
            ib_start_time = tz.localize(datetime(first_date.year, first_date.month, first_date.day,
                                                  ib_start_hour, ib_start_min))
            ib_end_time = tz.localize(datetime(first_date.year, first_date.month, first_date.day,
                                                ib_end_hour, ib_end_min))
            ib_start_num = mdates.date2num(ib_start_time)
            ib_end_num = mdates.date2num(ib_end_time)

            # IB rectangle (bounded box for IB period)
            ib_rect = Rectangle(
                (ib_start_num, ibl),
                ib_end_num - ib_start_num,
                ibh - ibl,
                facecolor="purple", edgecolor="purple", alpha=0.15, linewidth=2
            )
            ax.add_patch(ib_rect)

            # Extended IB levels (semi-transparent, after IB period only)
            ax.hlines(y=ibh, xmin=ib_end_num, xmax=chart_end_num, color="purple",
                     linestyle="--", linewidth=1, alpha=0.4, label=f"IBH: {ibh:.2f}")
            ax.hlines(y=ibl, xmin=ib_end_num, xmax=chart_end_num, color="purple",
                     linestyle="--", linewidth=1, alpha=0.4, label=f"IBL: {ibl:.2f}")
            ax.hlines(y=eq, xmin=ib_end_num, xmax=chart_end_num, color="orange",
                     linestyle="-.", linewidth=1, alpha=0.5, label=f"EQ: {eq:.2f}")

        # Plot signals
        for sig in signals:
            sig_time = mdates.date2num(sig["time"].astimezone(tz))
            sig_price = sig["entry_price"]
            marker = "^" if sig["direction"] == "long" else "v"
            color = "blue"
            ax.scatter(sig_time, sig_price, marker=marker, s=200, c=color, zorder=5,
                      label=f"Signal: {sig['variation']} {sig['direction']}")

        # Plot trade entry/exit
        if trade:
            entry_time = mdates.date2num(trade.entry_time.astimezone(tz))
            ax.axvline(x=entry_time, color="blue", linestyle=":", alpha=0.7)
            ax.scatter(entry_time, trade.entry_price, marker="o", s=150, c="blue",
                      edgecolors="black", zorder=5, label=f"Entry: {trade.entry_price:.2f}")

            if trade.exit_time:
                exit_time = mdates.date2num(trade.exit_time.astimezone(tz))
                ax.axvline(x=exit_time, color="gray", linestyle=":", alpha=0.5)
                ax.scatter(exit_time, trade.exit_price, marker="x", s=150, c="red",
                          zorder=5, label=f"Exit: {trade.exit_price:.2f}")

            # Always draw initial SL/TP as extended reference lines
            entry_time_num = mdates.date2num(trade.entry_time.astimezone(tz))
            exit_time_num = mdates.date2num(trade.exit_time.astimezone(tz)) if trade.exit_time else chart_end_num

            if trade.sl:
                # Initial SL - extended dashed line
                ax.hlines(y=trade.sl, xmin=entry_time_num, xmax=chart_end_num,
                         color="red", linestyle="--", linewidth=1.0, alpha=0.5,
                         label=f"SL: {trade.sl:.2f}")
            if trade.tp:
                # Initial TP - extended dashed line
                ax.hlines(y=trade.tp, xmin=entry_time_num, xmax=chart_end_num,
                         color="green", linestyle="--", linewidth=1.0, alpha=0.5,
                         label=f"TP: {trade.tp:.2f}")

            # TSL levels as "staircase" - shows how SL/TP moved over time
            # Use tsl_state from result (preserved before clearing) or fall back to strategy
            tsl_state = result.get("tsl_state") or self.strategy.tsl_state
            if tsl_state and trade.sl and trade.tp and tsl_state.get("tsl_history"):
                tsl_history = tsl_state.get("tsl_history", [])
                entry_time_num = mdates.date2num(trade.entry_time.astimezone(tz))
                exit_time_num = mdates.date2num(trade.exit_time.astimezone(tz)) if trade.exit_time else chart_end_num

                # Build list of TSL segments: [(start_time, end_time, sl, tp), ...]
                segments = []

                # Initial segment: from entry to first TSL change (or exit if no changes)
                initial_sl = trade.sl
                initial_tp = trade.tp

                if tsl_history:
                    # First segment: entry -> first TSL change
                    first_change_time = mdates.date2num(tsl_history[0]["time"].astimezone(tz))
                    segments.append((entry_time_num, first_change_time, initial_sl, initial_tp))

                    # Middle segments: each TSL change to the next
                    for i, change in enumerate(tsl_history):
                        change_time = mdates.date2num(change["time"].astimezone(tz))
                        if i + 1 < len(tsl_history):
                            next_time = mdates.date2num(tsl_history[i + 1]["time"].astimezone(tz))
                        else:
                            next_time = exit_time_num
                        segments.append((change_time, next_time, change["sl"], change["tp"]))
                else:
                    # No TSL changes - single segment from entry to exit
                    segments.append((entry_time_num, exit_time_num, initial_sl, initial_tp))

                # Draw segments
                for i, (start, end, sl, tp) in enumerate(segments):
                    alpha = 0.3 if i == 0 else 0.5  # First segment fainter
                    # SL line (red/darkred)
                    ax.hlines(y=sl, xmin=start, xmax=end, color="darkred",
                             linestyle=":", linewidth=0.8, alpha=alpha)
                    # TP line (green/darkgreen)
                    ax.hlines(y=tp, xmin=start, xmax=end, color="darkgreen",
                             linestyle=":", linewidth=0.8, alpha=alpha)

                # Add legend entries for TSL (initial SL/TP already added above)
                if tsl_history:
                    final_sl = segments[-1][2]  # Last segment's SL
                    final_tp = segments[-1][3]  # Last segment's TP
                    ax.plot([], [], color="darkred", linestyle=":", linewidth=1.5,
                           alpha=0.8, label=f"TSL SL: {final_sl:.2f}")
                    ax.plot([], [], color="darkgreen", linestyle=":", linewidth=1.5,
                           alpha=0.8, label=f"TSL TP: {final_tp:.2f}")
                    ax.plot([], [], color="gray", linestyle=":", linewidth=1.5,
                           alpha=0.7, label=f"TSL steps: {len(tsl_history)}")

        # Format - anchor ticks at :00 and :30
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M", tz=tz))
        ax.xaxis.set_major_locator(mdates.MinuteLocator(byminute=[0, 30]))
        plt.xticks(rotation=45)

        # Title - calculate R based on actual dollar risk, not price distance
        if trade:
            profit_str = f"+${trade.profit:.2f}" if trade.profit > 0 else f"-${abs(trade.profit):.2f}"
            # R based on dollar risk (same as backtest)
            if self.risk_amount:
                risk_money = self.risk_amount
            elif self.risk_pct:
                risk_money = self.initial_balance * self.risk_pct / 100
            else:
                risk_money = abs(trade.entry_price - trade.sl) if trade.sl else 1  # fallback
            r_value = trade.profit / risk_money if risk_money > 0 else 0
            title = f"{self.symbol} - {trade.entry_time.strftime('%Y-%m-%d')} | " \
                    f"{trade.variation} {trade.direction.upper()} | " \
                    f"P/L: {profit_str} | R: {r_value:+.2f} | Exit: {trade.exit_reason}"
        else:
            date_str = candles["time"].iloc[0].strftime("%Y-%m-%d")
            title = f"{self.symbol} - {date_str} | No trade"

        ax.set_title(title, fontsize=12, fontweight="bold")
        ax.set_xlabel("Time")
        ax.set_ylabel("Price")
        ax.legend(loc="upper left", fontsize=8)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        return fig


def export_trades_for_debug(
    trade_log: List[TradeLog],
    output_path: str,
    ib_data_by_date: Optional[Dict[str, Dict]] = None,
) -> str:
    """
    Export trades to Excel with columns for manual comments.

    Args:
        trade_log: List of TradeLog objects
        output_path: Output Excel file path
        ib_data_by_date: Optional dict of IB data by date string

    Returns:
        Path to created file
    """
    if not trade_log:
        print("No trades to export")
        return None

    rows = []
    for i, trade in enumerate(trade_log, 1):
        date_str = trade.entry_time.strftime("%Y-%m-%d") if trade.entry_time else ""

        # Get IB data if available
        ib_data = {}
        if ib_data_by_date and date_str in ib_data_by_date:
            ib_data = ib_data_by_date[date_str]

        # Calculate R
        if trade.sl and trade.entry_price:
            risk = abs(trade.entry_price - trade.sl)
            r_value = trade.profit / risk if risk > 0 else 0
        else:
            r_value = 0

        rows.append({
            "#": i,
            "Date": date_str,
            "Symbol": trade.symbol,
            "Variation": trade.variation or "Unknown",
            "Direction": trade.direction.upper() if trade.direction else "",
            "IBH": ib_data.get("ibh", ""),
            "IBL": ib_data.get("ibl", ""),
            "EQ": ib_data.get("eq", ""),
            "Entry Time": trade.entry_time.strftime("%H:%M:%S") if trade.entry_time else "",
            "Entry Price": trade.entry_price,
            "Stop Loss": trade.sl,
            "Take Profit": trade.tp,
            "Exit Time": trade.exit_time.strftime("%H:%M:%S") if trade.exit_time else "",
            "Exit Price": trade.exit_price,
            "Exit Reason": trade.exit_reason or "",
            "P/L ($)": trade.profit,
            "R": round(r_value, 2),
            "Status": "",  # For manual input: OK, ISSUE, SKIP
            "Comment": "",  # For manual notes
        })

    df = pd.DataFrame(rows)

    # Export to Excel with formatting
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"

        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Color P/L column
                if headers[col_idx-1] == "P/L ($)" and isinstance(value, (int, float)):
                    if value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Set column widths
        col_widths = {
            "A": 5, "B": 12, "C": 8, "D": 12, "E": 10,
            "F": 12, "G": 12, "H": 12, "I": 10, "J": 12,
            "K": 12, "L": 12, "M": 10, "N": 12, "O": 10,
            "P": 10, "Q": 8, "R": 10, "S": 40,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)
        print(f"Exported {len(trade_log)} trades to {output_path}")
        return output_path

    except ImportError:
        # Fallback to pandas
        df.to_excel(output_path, index=False)
        print(f"Exported {len(trade_log)} trades to {output_path}")
        return output_path


def load_reference_trades(xlsx_path: str, symbol: str = "GER40") -> pd.DataFrame:
    """
    Load reference trades from the dual_asset Excel file.

    Args:
        xlsx_path: Path to reference Excel file
        symbol: Filter by symbol (GER40 or XAUUSD)

    Returns:
        DataFrame with reference trades
    """
    df = pd.read_excel(xlsx_path, sheet_name=0)

    # Rename columns (they're in Russian)
    column_map = {
        df.columns[0]: "Date",
        df.columns[1]: "Symbol",
        df.columns[2]: "Status",
        df.columns[3]: "IB_H",
        df.columns[4]: "IB_L",
        df.columns[5]: "EQ",
        df.columns[6]: "Variation",
        df.columns[7]: "Direction",
        df.columns[8]: "IB_Params",
        df.columns[9]: "Entry_Time",
        df.columns[10]: "Entry_Price",
        df.columns[11]: "SL",
        df.columns[12]: "TP",
        df.columns[13]: "SL_Adjusted",
        df.columns[14]: "Exit_Time",
        df.columns[15]: "Exit_Price",
        df.columns[16]: "Exit_Reason",
        df.columns[17]: "R",
    }
    df = df.rename(columns=column_map)

    # Filter by symbol
    if symbol:
        df = df[df["Symbol"] == symbol].copy()

    return df
