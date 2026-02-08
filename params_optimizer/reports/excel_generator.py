"""
Excel Report Generator for Parameter Optimization Results.

Generates formatted Excel reports similar to the old anal.py output,
but from parquet data instead of thousands of Excel files.
"""

import json
from pathlib import Path
from typing import Dict, List, Optional, Any

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Style constants
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
SILVER_FILL = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
BRONZE_FILL = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Variation colors
VARIATION_COLORS = {
    "OCAE": "4472C4",     # Blue
    "TCWE": "70AD47",     # Green
    "Reverse": "FFC000",  # Gold
    "REV_RB": "ED7D31",   # Orange
}


class ExcelReportGenerator:
    """
    Generates Excel reports from optimization results.

    Similar output to anal.py but from parquet instead of Excel files.
    """

    def __init__(self, output_dir: Path):
        """
        Initialize generator.

        Args:
            output_dir: Directory for output files
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        results_path: Path,
        symbol: str,
        top_n: int = 500,
    ) -> Path:
        """
        Generate full Excel report from parquet results.

        Args:
            results_path: Path to results parquet file
            symbol: Trading symbol (for filename)
            top_n: Number of top results per variation sheet

        Returns:
            Path to generated Excel file
        """
        # Load results
        df = pd.read_parquet(results_path)

        # Parse params if stored as JSON
        if "params_json" in df.columns:
            df["params"] = df["params_json"].apply(json.loads)
            df = df.drop(columns=["params_json"])

        # Flatten params to columns
        df = self._flatten_params(df)

        # Create workbook
        wb = Workbook()

        # Sheet 1: GridResults (all combinations)
        self._create_grid_sheet(wb, df, "GridResults")

        # Sheet 2: Top100
        top100 = df.nlargest(100, "total_r")
        self._create_grid_sheet(wb, top100, "Top100", highlight_top3=True)

        # Sheets 3-6: Variation sheets
        for var in ["OCAE", "TCWE", "Reverse", "REV_RB"]:
            var_col = f"{var.lower()}_total_r"
            if var_col in df.columns:
                # Sort by variation's total_r and take top N
                var_df = df[df[f"{var.lower()}_trades"] > 0].nlargest(top_n, var_col)
                self._create_variation_sheet(wb, var_df, var, var.lower())

        # Sheet 7: Summary
        self._create_summary_sheet(wb, df)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Save
        output_path = self.output_dir / f"{symbol}_analysis.xlsx"
        wb.save(output_path)

        return output_path

    def _flatten_params(self, df: pd.DataFrame) -> pd.DataFrame:
        """Flatten params dict to separate columns."""
        if "params" not in df.columns:
            return df

        result = df.copy()

        # Get param keys from first non-null params
        sample_params = None
        for idx, row in result.iterrows():
            if isinstance(row.get("params"), dict):
                sample_params = row["params"]
                break

        if sample_params is None:
            return result

        # Extract each param to column
        for key in sample_params.keys():
            result[f"param_{key}"] = result["params"].apply(
                lambda p: p.get(key) if isinstance(p, dict) else None
            )

        # Drop original params column
        result = result.drop(columns=["params"])

        return result

    def _create_grid_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        sheet_name: str,
        highlight_top3: bool = False
    ) -> None:
        """Create grid results sheet."""
        ws = wb.create_sheet(sheet_name)

        # Select columns to display
        display_cols = self._get_display_columns(df)
        display_df = df[display_cols].copy()

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

                # Header formatting
                if r_idx == 0:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    # Alternating row colors
                    if r_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

                    # Top 3 highlighting
                    if highlight_top3 and r_idx <= 3:
                        if r_idx == 1:
                            cell.fill = GOLD_FILL
                        elif r_idx == 2:
                            cell.fill = SILVER_FILL
                        elif r_idx == 3:
                            cell.fill = BRONZE_FILL

                cell.border = THIN_BORDER

        # Auto-width columns
        self._auto_width_columns(ws, max_width=50)

        # Conditional formatting for total_r column (only if we have data)
        if "total_r" in display_cols and len(df) > 0:
            col_idx = display_cols.index("total_r") + 1
            col_letter = get_column_letter(col_idx)
            self._add_color_scale(ws, col_letter, 2, len(df) + 1)

        # Conditional formatting for winrate column (only if we have data)
        if "winrate" in display_cols and len(df) > 0:
            col_idx = display_cols.index("winrate") + 1
            col_letter = get_column_letter(col_idx)
            self._add_color_scale(ws, col_letter, 2, len(df) + 1, mid_value=50)

    def _create_variation_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        var_name: str,
        var_prefix: str
    ) -> None:
        """Create variation-specific sheet."""
        ws = wb.create_sheet(var_name)

        # Get variation-specific columns + params
        var_cols = [c for c in df.columns if c.startswith(f"{var_prefix}_")]
        param_cols = [c for c in df.columns if c.startswith("param_")]

        # Rename var columns to remove prefix for cleaner display
        display_df = df[var_cols + param_cols].copy()
        display_df.columns = [
            c.replace(f"{var_prefix}_", "") if c.startswith(f"{var_prefix}_") else c
            for c in display_df.columns
        ]

        # Reorder: metrics first, then params
        metric_cols = [c for c in display_df.columns if not c.startswith("param_")]
        param_cols = [c for c in display_df.columns if c.startswith("param_")]
        display_df = display_df[metric_cols + param_cols]

        # Get variation color
        var_color = VARIATION_COLORS.get(var_name, "366092")
        header_fill = PatternFill(start_color=var_color, end_color=var_color, fill_type="solid")

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

                if r_idx == 0:
                    cell.fill = header_fill
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    if r_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL

                cell.border = THIN_BORDER

        self._auto_width_columns(ws, max_width=50)

        # Conditional formatting for total_r (only if we have data)
        if "total_r" in display_df.columns and len(df) > 0:
            col_idx = list(display_df.columns).index("total_r") + 1
            col_letter = get_column_letter(col_idx)
            self._add_color_scale(ws, col_letter, 2, len(df) + 1)

    def _create_summary_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create summary sheet comparing variations."""
        ws = wb.create_sheet("Summary")

        # Calculate summary stats for each variation
        variations = ["OCAE", "TCWE", "Reverse", "REV_RB"]
        summary_data = []

        for var in variations:
            prefix = var.lower()
            trades_col = f"{prefix}_trades"
            total_r_col = f"{prefix}_total_r"
            winrate_col = f"{prefix}_winrate"
            sharpe_col = f"{prefix}_sharpe_ratio"

            if trades_col not in df.columns:
                continue

            # Filter configs with trades for this variation
            var_df = df[df[trades_col] > 0]

            if len(var_df) == 0:
                summary_data.append({
                    "Variation": var,
                    "Total_configs": 0,
                    "Total_trades": 0,
                    "Avg_Total_R": 0,
                    "Max_Total_R": 0,
                    "Min_Total_R": 0,
                    "Avg_Winrate": 0,
                    "Max_Winrate": 0,
                    "Avg_Sharpe": 0,
                    "Best_config_R": 0,
                })
                continue

            # Best config (highest total_r for this variation)
            best_idx = var_df[total_r_col].idxmax()
            best_r = var_df.loc[best_idx, total_r_col]

            summary_data.append({
                "Variation": var,
                "Total_configs": len(var_df),
                "Total_trades": int(var_df[trades_col].sum()),
                "Avg_Total_R": round(var_df[total_r_col].mean(), 2),
                "Max_Total_R": round(var_df[total_r_col].max(), 2),
                "Min_Total_R": round(var_df[total_r_col].min(), 2),
                "Avg_Winrate": round(var_df[winrate_col].mean(), 2),
                "Max_Winrate": round(var_df[winrate_col].max(), 2),
                "Avg_Sharpe": round(var_df[sharpe_col].mean(), 4) if sharpe_col in df.columns else 0,
                "Best_config_R": round(best_r, 2),
            })

        summary_df = pd.DataFrame(summary_data)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True)):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

                if r_idx == 0:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center")
                else:
                    # Color by variation
                    var_name = summary_df.iloc[r_idx - 1]["Variation"]
                    var_color = VARIATION_COLORS.get(var_name, "FFFFFF")
                    if c_idx == 1:  # Variation name column
                        cell.fill = PatternFill(start_color=var_color, end_color=var_color, fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF" if var_name != "Reverse" else "000000")

                cell.border = THIN_BORDER

        self._auto_width_columns(ws, max_width=30)

    def _get_display_columns(self, df: pd.DataFrame) -> List[str]:
        """Get columns to display in grid sheet (ordered)."""
        # Primary metrics
        primary = ["total_r", "total_trades", "winrate", "sharpe_ratio", "profit_factor", "max_drawdown"]

        # Variation totals
        var_totals = []
        for var in ["ocae", "tcwe", "reverse", "rev_rb"]:
            for metric in ["total_r", "trades", "winrate"]:
                col = f"{var}_{metric}"
                if col in df.columns:
                    var_totals.append(col)

        # Parameters
        params = sorted([c for c in df.columns if c.startswith("param_")])

        # Combine (only existing columns)
        result = []
        for col in primary + var_totals + params:
            if col in df.columns:
                result.append(col)

        return result

    def _auto_width_columns(self, ws, max_width: int = 50) -> None:
        """Auto-fit column widths."""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells[:100]:  # Check first 100 rows only
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column].width = adjusted_width

    def _add_color_scale(
        self,
        ws,
        column: str,
        start_row: int,
        end_row: int,
        mid_value: Optional[float] = None
    ) -> None:
        """Add color scale conditional formatting."""
        cell_range = f"{column}{start_row}:{column}{end_row}"

        if mid_value is not None:
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num", mid_value=mid_value, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
        else:
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )

        ws.conditional_formatting.add(cell_range, rule)


def generate_excel_report(
    results_path: Path,
    output_dir: Path,
    symbol: str,
    top_n: int = 500
) -> Path:
    """
    Convenience function to generate Excel report.

    Args:
        results_path: Path to parquet results
        output_dir: Output directory
        symbol: Trading symbol
        top_n: Top N results per variation

    Returns:
        Path to generated Excel file
    """
    generator = ExcelReportGenerator(output_dir)
    return generator.generate_report(results_path, symbol, top_n)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate Excel report from optimization results")
    parser.add_argument("--results", type=Path, required=True, help="Path to results parquet")
    parser.add_argument("--output", type=Path, default=Path("."), help="Output directory")
    parser.add_argument("--symbol", type=str, required=True, help="Trading symbol")
    parser.add_argument("--top-n", type=int, default=500, help="Top N per variation")

    args = parser.parse_args()

    output_path = generate_excel_report(args.results, args.output, args.symbol, args.top_n)
    print(f"Report generated: {output_path}")
